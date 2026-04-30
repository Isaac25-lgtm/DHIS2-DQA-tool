from __future__ import annotations

from io import BytesIO
from uuid import UUID

from fastapi import HTTPException, status
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload, joinedload

from app.models.assessment_facility import AssessmentFacility
from app.models.assessment_facility_team_member import AssessmentFacilityTeamMember
from app.models.assessment_round import AssessmentRound
from app.models.assessment_round_indicator import AssessmentRoundIndicator
from app.models.base import AssessmentFacilityStatus, AssessmentTeamRole, SeverityLevel
from app.models.dqa_value import DqaValue
from app.schemas.submissions import (
    SubmissionDashboardResponse,
    SubmissionDetailResponse,
    SubmissionListItemResponse,
    SubmissionStatsResponse,
    SubmissionTeamLeadOptionResponse,
    SubmissionValueRowResponse,
)
from app.services.scoring_service import calculate_facility_score

SUBMITTED_STATUSES = {
    AssessmentFacilityStatus.SUBMITTED,
    AssessmentFacilityStatus.UNDER_REVIEW,
    AssessmentFacilityStatus.APPROVED,
    AssessmentFacilityStatus.CLOSED,
}


def _assessment_facility_query():
    return (
        select(AssessmentFacility)
        .options(
            joinedload(AssessmentFacility.facility),
            joinedload(AssessmentFacility.assessment_round).selectinload(AssessmentRound.selected_indicators).joinedload(AssessmentRoundIndicator.indicator),
            selectinload(AssessmentFacility.team_members).joinedload(AssessmentFacilityTeamMember.user),
            selectinload(AssessmentFacility.dqa_values).joinedload(DqaValue.indicator),
        )
    )


def _list_assessment_facilities(db: Session, assessment_round_id: UUID | None = None) -> list[AssessmentFacility]:
    statement = _assessment_facility_query()
    if assessment_round_id:
        statement = statement.where(AssessmentFacility.assessment_round_id == assessment_round_id)
    return list(db.scalars(statement).unique())


def _filter_by_assessment_facility_id(items: list[AssessmentFacility], assessment_facility_id: UUID | None = None) -> list[AssessmentFacility]:
    if not assessment_facility_id:
        return items
    return [item for item in items if item.id == assessment_facility_id]


def _selected_indicator_map(assessment_facility: AssessmentFacility) -> dict[UUID, AssessmentRoundIndicator]:
    return {
        item.indicator_id: item
        for item in sorted(assessment_facility.assessment_round.selected_indicators, key=lambda item: item.display_order)
    }


def _team_lead_member(assessment_facility: AssessmentFacility) -> AssessmentFacilityTeamMember | None:
    active_members = [item for item in assessment_facility.team_members if item.is_active]
    return next((item for item in active_members if item.team_role == AssessmentTeamRole.TEAM_LEAD), None)


def _team_names(assessment_facility: AssessmentFacility) -> tuple[UUID | None, str | None, list[str]]:
    active_members = [item for item in assessment_facility.team_members if item.is_active]
    lead_member = _team_lead_member(assessment_facility)
    lead_id = lead_member.user_id if lead_member else None
    lead = lead_member.user.full_name if lead_member else None
    members = [
        item.user.full_name
        for item in active_members
        if item.team_role == AssessmentTeamRole.TEAM_MEMBER
    ]
    return lead_id, lead, members


def _filter_by_team_lead(items: list[AssessmentFacility], team_lead_user_id: UUID | None = None) -> list[AssessmentFacility]:
    if not team_lead_user_id:
        return items
    return [
        item
        for item in items
        if (lead_member := _team_lead_member(item)) is not None and lead_member.user_id == team_lead_user_id
    ]


def _team_lead_options(items: list[AssessmentFacility]) -> list[SubmissionTeamLeadOptionResponse]:
    options: dict[UUID, str] = {}
    for item in items:
        lead_member = _team_lead_member(item)
        if lead_member:
            options[lead_member.user_id] = lead_member.user.full_name
    return [
        SubmissionTeamLeadOptionResponse(user_id=user_id, full_name=full_name)
        for user_id, full_name in sorted(options.items(), key=lambda entry: entry[1].lower())
    ]


def _percent_diff(reference_value: int | None, comparison_value: int | None) -> float | None:
    if reference_value is None or comparison_value is None:
        return None
    if reference_value == 0 and comparison_value == 0:
        return 0.0
    if reference_value == 0:
        return None
    return round(abs(comparison_value - reference_value) / abs(reference_value) * 100, 2)


def _value_percentages(value: DqaValue | None) -> tuple[float | None, float | None, float | None, float | None]:
    if value is None:
        return None, None, None, None
    register_hmis = _percent_diff(value.register_value, value.hmis105_value)
    hmis_dhis2 = _percent_diff(value.hmis105_value, value.dhis2_value_at_assessment)
    register_dhis2 = _percent_diff(value.register_value, value.dhis2_value_at_assessment)
    valid = [item for item in (register_hmis, hmis_dhis2, register_dhis2) if item is not None]
    return register_hmis, hmis_dhis2, register_dhis2, max(valid) if valid else None


def _live_flag_for_value(value: DqaValue | None) -> str:
    if value is None or value.register_value is None or value.hmis105_value is None or value.dhis2_value_at_assessment is None:
        return "Incomplete"
    if value.indicator and value.indicator.is_death_indicator:
        values = [value.register_value, value.hmis105_value, value.dhis2_value_at_assessment]
        if max(values) - min(values) >= 1:
            return "Critical"
    _, _, _, max_percent = _value_percentages(value)
    if max_percent is None:
        return "Flagged >5%"
    if max_percent == 0:
        return "Match"
    if max_percent <= 5:
        return "Within 5%"
    return "Flagged >5%"


def _flag_for_value(value: DqaValue | None) -> str:
    if value is None or value.severity is None:
        return _live_flag_for_value(value)
    if value.severity == SeverityLevel.EXACT:
        return "Match"
    if value.severity == SeverityLevel.MINOR:
        return "Within 5%"
    if value.severity in {SeverityLevel.MAJOR, SeverityLevel.MODERATE}:
        return "Flagged >5%"
    if value.severity == SeverityLevel.CRITICAL:
        return "Critical"
    if value.severity == SeverityLevel.MISSING:
        return "Incomplete"
    return value.severity.value


def _score_for_facility(assessment_facility: AssessmentFacility) -> dict:
    required_ids = {item.indicator_id for item in assessment_facility.assessment_round.selected_indicators if item.is_required}
    return calculate_facility_score(
        assessment_facility.dqa_values,
        required_ids,
        assessment_facility.assessment_round.scoring_settings_json,
    )


def serialize_submission_item(assessment_facility: AssessmentFacility) -> SubmissionListItemResponse:
    selected_indicator_ids = {item.indicator_id for item in assessment_facility.assessment_round.selected_indicators}
    values = [item for item in assessment_facility.dqa_values if item.indicator_id in selected_indicator_ids]
    lead_id, lead, members = _team_names(assessment_facility)
    completed = sum(1 for item in values if item.register_value is not None and item.hmis105_value is not None)
    flagged = sum(1 for item in values if item.severity in {SeverityLevel.MODERATE, SeverityLevel.MAJOR, SeverityLevel.CRITICAL})
    critical = sum(1 for item in values if item.severity == SeverityLevel.CRITICAL)
    score = _score_for_facility(assessment_facility)
    last_synced = max((item.last_synced_at for item in values if item.last_synced_at), default=None)

    return SubmissionListItemResponse(
        assessment_facility_id=assessment_facility.id,
        assessment_round_id=assessment_facility.assessment_round_id,
        assessment_round_name=assessment_facility.assessment_round.name,
        reporting_period=assessment_facility.assessment_round.reporting_period,
        facility_id=assessment_facility.facility_id,
        facility_name=assessment_facility.facility.facility_name,
        district=assessment_facility.facility.district,
        status=assessment_facility.status.value,
        team_lead_user_id=lead_id,
        team_lead=lead,
        team_members=members,
        submitted_at=assessment_facility.submitted_at,
        last_synced_at=last_synced,
        completed_indicators=completed,
        total_indicators=len(selected_indicator_ids),
        flagged_rows=flagged,
        critical_rows=critical,
        dqa_score=float(score["score_percent"]),
        score_category=str(score["score_category"]),
        general_assessment_comment=assessment_facility.general_assessment_comment,
    )


def _serialize_submission_rows(assessment_facility: AssessmentFacility) -> list[SubmissionValueRowResponse]:
    values_by_indicator = {item.indicator_id: item for item in assessment_facility.dqa_values}
    rows: list[SubmissionValueRowResponse] = []
    for selected in sorted(assessment_facility.assessment_round.selected_indicators, key=lambda item: item.display_order):
        value = values_by_indicator.get(selected.indicator_id)
        register_hmis_percent_diff, hmis_dhis2_percent_diff, register_dhis2_percent_diff, max_percent_diff = _value_percentages(value)
        rows.append(
            SubmissionValueRowResponse(
                dqa_value_id=value.id if value else None,
                indicator_id=selected.indicator_id,
                indicator_name=selected.indicator.indicator_name,
                hmis_code=selected.indicator.hmis_code,
                source_register=selected.indicator.source_register,
                register_value=value.register_value if value else None,
                hmis105_value=value.hmis105_value if value else None,
                dhis2_value_at_assessment=value.dhis2_value_at_assessment if value else None,
                register_vs_hmis_difference=value.register_vs_hmis_difference if value else None,
                hmis_vs_dhis2_difference=value.hmis_vs_dhis2_difference if value else None,
                register_vs_dhis2_difference=value.register_vs_dhis2_difference if value else None,
                register_hmis_percent_diff=register_hmis_percent_diff,
                hmis_dhis2_percent_diff=hmis_dhis2_percent_diff,
                register_dhis2_percent_diff=register_dhis2_percent_diff,
                max_percent_diff=max_percent_diff,
                discrepancy_percent=float(value.discrepancy_percent) if value and value.discrepancy_percent is not None else None,
                issue_type=value.issue_type.value if value and value.issue_type else None,
                severity=value.severity.value if value and value.severity else None,
                flag=_flag_for_value(value),
                comparison_notes=value.comparison_notes if value else None,
                assessor_comment=value.assessor_comment if value else None,
                manager_comment=value.manager_comment if value else None,
            )
        )
    return rows


def build_submission_stats(items: list[AssessmentFacility]) -> SubmissionStatsResponse:
    submitted_items = [item for item in items if item.status in SUBMITTED_STATUSES]
    values = [value for item in submitted_items for value in item.dqa_values]
    scores = [_score_for_facility(item) for item in submitted_items]
    submitted_count = len(submitted_items)
    total = len(items)
    in_progress_count = sum(1 for item in items if item.status in {AssessmentFacilityStatus.IN_PROGRESS, AssessmentFacilityStatus.DRAFT_SAVED, AssessmentFacilityStatus.PENDING_SYNC})
    not_started_count = sum(1 for item in items if item.status in {AssessmentFacilityStatus.NOT_STARTED, AssessmentFacilityStatus.ASSIGNED})
    exact = sum(1 for item in values if item.severity == SeverityLevel.EXACT)
    within = sum(1 for item in values if item.severity == SeverityLevel.MINOR)
    flagged = sum(1 for item in values if item.severity in {SeverityLevel.MODERATE, SeverityLevel.MAJOR, SeverityLevel.CRITICAL})
    critical = sum(1 for item in values if item.severity == SeverityLevel.CRITICAL)
    missing = sum(1 for item in values if item.severity == SeverityLevel.MISSING)
    completion = round((submitted_count / total) * 100, 2) if total else 0.0
    average_score = round(sum(float(item["score_percent"]) for item in scores) / len(scores), 2) if scores else 0.0
    return SubmissionStatsResponse(
        total_facilities=total,
        submitted_facilities=submitted_count,
        pending_facilities=max(total - submitted_count, 0),
        in_progress_facilities=in_progress_count,
        not_started_facilities=not_started_count,
        completion_percent=completion,
        remaining_percent=round(max(100 - completion, 0), 2),
        total_submitted_rows=len(values),
        exact_count=exact,
        within_threshold_count=within,
        flagged_count=flagged,
        critical_count=critical,
        missing_count=missing,
        average_score_percent=average_score,
    )


def get_submissions_dashboard(
    db: Session,
    assessment_round_id: UUID | None = None,
    team_lead_user_id: UUID | None = None,
) -> SubmissionDashboardResponse:
    facilities = _list_assessment_facilities(db, assessment_round_id)
    filtered_facilities = _filter_by_team_lead(facilities, team_lead_user_id)
    submitted = [item for item in filtered_facilities if item.status in SUBMITTED_STATUSES]
    return SubmissionDashboardResponse(
        stats=build_submission_stats(filtered_facilities),
        team_leads=_team_lead_options(facilities),
        submissions=[serialize_submission_item(item) for item in sorted(submitted, key=lambda item: item.submitted_at or item.updated_at, reverse=True)],
    )


def get_submission_detail(db: Session, assessment_facility_id: UUID) -> SubmissionDetailResponse:
    assessment_facility = db.execute(
        _assessment_facility_query().where(AssessmentFacility.id == assessment_facility_id)
    ).unique().scalar_one_or_none()
    if not assessment_facility:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Submission not found.")
    return SubmissionDetailResponse(
        summary=serialize_submission_item(assessment_facility),
        values=_serialize_submission_rows(assessment_facility),
    )


def _style_header(sheet) -> None:
    fill = PatternFill("solid", fgColor="0F4C81")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _flag_fill(flag: str | None) -> PatternFill | None:
    colors = {
        "Match": "C6EFCE",
        "Within 5%": "DDEBF7",
        "Flagged >5%": "FCE4D6",
        "Critical": "FFC7CE",
        "Incomplete": "D9EAD3",
        "EXACT": "C6EFCE",
        "MINOR": "DDEBF7",
        "MODERATE": "FFF2CC",
        "MAJOR": "FCE4D6",
        "MISSING": "D9EAD3",
        "CRITICAL": "FFC7CE",
    }
    color = colors.get(flag or "")
    return PatternFill("solid", fgColor=color) if color else None


def _style_body(sheet) -> None:
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        if row[0].row % 2 == 0:
            for cell in row:
                if cell.fill.fill_type is None:
                    cell.fill = PatternFill("solid", fgColor="F8FBFD")


def _finish_sheet(sheet) -> None:
    _style_header(sheet)
    _style_body(sheet)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    _autosize(sheet)


def _autosize(sheet) -> None:
    for column in sheet.columns:
        letter = column[0].column_letter
        width = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[letter].width = min(max(width + 2, 12), 55)


def build_submissions_workbook(
    db: Session,
    assessment_round_id: UUID | None = None,
    assessment_facility_id: UUID | None = None,
    team_lead_user_id: UUID | None = None,
) -> bytes:
    if assessment_facility_id:
        detail = get_submission_detail(db, assessment_facility_id)
        items = [detail.summary]
        rows_by_facility = {assessment_facility_id: detail.values}
        stats = build_submission_stats(
            _filter_by_assessment_facility_id(_list_assessment_facilities(db, detail.summary.assessment_round_id), assessment_facility_id)
        )
    else:
        facilities = _filter_by_team_lead(_list_assessment_facilities(db, assessment_round_id), team_lead_user_id)
        facilities = _filter_by_assessment_facility_id(facilities, assessment_facility_id)
        submitted = [item for item in facilities if item.status in SUBMITTED_STATUSES]
        items = [serialize_submission_item(item) for item in sorted(submitted, key=lambda item: item.submitted_at or item.updated_at, reverse=True)]
        rows_by_facility = {item.assessment_facility_id: get_submission_detail(db, item.assessment_facility_id).values for item in items}
        stats = build_submission_stats(facilities)

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary.append(["Metric", "Value"])
    for key, value in stats.model_dump().items():
        summary.append([key.replace("_", " ").title(), value])
    _finish_sheet(summary)

    submissions = workbook.create_sheet("Submissions")
    submissions.append([
        "Assessment Round",
        "Reporting Period",
        "Facility",
        "District",
        "Status",
        "Group Account",
        "Submitted At",
        "DQA Score",
        "Score Category",
        "Completed Indicators",
        "Total Indicators",
        "Flagged Rows",
        "Critical Rows",
        "Team Members",
        "General Comment",
    ])
    for item in items:
        submissions.append([
            item.assessment_round_name,
            item.reporting_period,
            item.facility_name,
            item.district,
            item.status,
            item.team_lead,
            item.submitted_at.isoformat() if item.submitted_at else "",
            item.dqa_score,
            item.score_category,
            item.completed_indicators,
            item.total_indicators,
            item.flagged_rows,
            item.critical_rows,
            ", ".join(item.team_members),
            item.general_assessment_comment or "",
        ])
    _finish_sheet(submissions)

    data = workbook.create_sheet("Submitted Data")
    data.append([
        "Assessment Round",
        "Reporting Period",
        "Facility",
        "District",
        "Group Account",
        "HMIS Code",
        "Indicator",
        "Source Register",
        "Register Value",
        "HMIS 105 Value",
        "DHIS2 Value",
        "HMIS vs Register %",
        "DHIS2 vs HMIS %",
        "DHIS2 vs Register %",
        "Max % Difference",
        "Flag",
        "Severity",
        "Issue Type",
        "Notes",
        "Field Assessor Comment",
        "Manager Comment",
        "General Facility Comment",
    ])
    meta_by_facility = {item.assessment_facility_id: item for item in items}
    for assessment_facility_id_key, rows in rows_by_facility.items():
        item = meta_by_facility.get(assessment_facility_id_key)
        for row in rows:
            data.append([
                item.assessment_round_name if item else "",
                item.reporting_period if item else "",
                item.facility_name if item else "",
                item.district if item else "",
                item.team_lead if item else "",
                row.hmis_code,
                row.indicator_name,
                row.source_register or "",
                row.register_value,
                row.hmis105_value,
                row.dhis2_value_at_assessment,
                row.register_hmis_percent_diff,
                row.hmis_dhis2_percent_diff,
                row.register_dhis2_percent_diff,
                row.max_percent_diff,
                row.flag,
                row.severity or "",
                row.issue_type or "",
                row.comparison_notes or "",
                row.assessor_comment or "",
                row.manager_comment or "",
                item.general_assessment_comment if item and item.general_assessment_comment else "",
            ])
    flag_col = 16
    severity_col = 17
    percent_cols = ["L", "M", "N", "O"]
    for row in data.iter_rows(min_row=2):
        for col_idx in (flag_col, severity_col):
            fill = _flag_fill(str(row[col_idx - 1].value or ""))
            if fill:
                row[col_idx - 1].fill = fill
                row[col_idx - 1].font = Font(bold=True)
    if data.max_row >= 2:
        for col in percent_cols:
            data.conditional_formatting.add(
                f"{col}2:{col}{data.max_row}",
                ColorScaleRule(start_type="num", start_value=0, start_color="C6EFCE", mid_type="num", mid_value=5, mid_color="FFF2CC", end_type="num", end_value=20, end_color="FFC7CE"),
            )
    _finish_sheet(data)

    comments = workbook.create_sheet("Field Comments")
    comments.append(["Facility", "Group Account", "HMIS Code", "Indicator", "Comment Type", "Comment"])
    for item in items:
        if item.general_assessment_comment:
            comments.append([item.facility_name, item.team_lead or "", "", "", "General Facility Comment", item.general_assessment_comment])
        for row in rows_by_facility.get(item.assessment_facility_id, []):
            if row.assessor_comment:
                comments.append([item.facility_name, item.team_lead or "", row.hmis_code, row.indicator_name, "Field Assessor Comment", row.assessor_comment])
            if row.manager_comment:
                comments.append([item.facility_name, item.team_lead or "", row.hmis_code, row.indicator_name, "Manager Comment", row.manager_comment])
    _finish_sheet(comments)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
