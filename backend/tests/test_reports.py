from __future__ import annotations

from datetime import date, timedelta
from io import BytesIO
from uuid import UUID

from docx import Document
from openpyxl import load_workbook

from app.models.ai_generation_log import AiGenerationLog
from app.models.base import ReportStatus, ReportType
from app.models.dqa_value import DqaValue
from app.models.export_log import ExportLog
from app.models.report import Report
from app.services.export_service import export_report_docx, export_report_xlsx


def _create_published_assignment(client, manager_token: str, facility_id: str, indicator_id: str, assessor_id: str) -> tuple[str, str]:
    round_response = client.post(
        "/api/assessment-rounds",
        headers={"Authorization": f"Bearer {manager_token}"},
        json={
            "name": "Prompt 6 Round",
            "description": "Report fixture",
            "reporting_period": "2026-03",
            "period_type": "MONTHLY",
            "start_date": "2026-03-01",
            "end_date": "2026-03-31",
            "deadline": "2026-04-20",
            "notes": "Prompt 6 reporting test",
        },
    )
    assert round_response.status_code == 201
    round_id = round_response.json()["id"]

    client.put(
        f"/api/assessment-rounds/{round_id}/indicators",
        headers={"Authorization": f"Bearer {manager_token}"},
        json={"indicators": [{"indicator_id": indicator_id, "display_order": 1, "is_required": True}]},
    )
    facilities_response = client.put(
        f"/api/assessment-rounds/{round_id}/facilities",
        headers={"Authorization": f"Bearer {manager_token}"},
        json={"facility_ids": [facility_id]},
    )
    assessment_facility_id = facilities_response.json()[0]["id"]
    client.post(
        f"/api/assessment-rounds/{round_id}/assign",
        headers={"Authorization": f"Bearer {manager_token}"},
        json={"assignments": [{"facility_id": facility_id, "assessor_id": assessor_id}]},
    )
    client.post(
        f"/api/assessment-rounds/{round_id}/publish",
        headers={"Authorization": f"Bearer {manager_token}"},
        json={"allow_unassigned_facilities": False},
    )
    return round_id, assessment_facility_id


def _prepare_compared_assessment(
    client,
    db_session,
    manager_token: str,
    assessor_token: str,
    active_facility,
    active_indicator,
    seeded_assessor,
) -> tuple[str, str]:
    round_id, assessment_facility_id = _create_published_assignment(
        client, manager_token, str(active_facility.id), str(active_indicator.id), str(seeded_assessor.id)
    )
    client.post(
        f"/api/my-assessments/{assessment_facility_id}/values",
        headers={"Authorization": f"Bearer {assessor_token}"},
        json={
            "values": [
                {
                    "indicator_id": str(active_indicator.id),
                    "register_value": 100,
                    "hmis105_value": 95,
                    "assessor_comment": "Assessor note should be optional in reports.",
                }
            ]
        },
    )
    saved_value = (
        db_session.query(DqaValue)
        .filter_by(assessment_facility_id=UUID(assessment_facility_id), indicator_id=active_indicator.id)
        .one()
    )
    saved_value.dhis2_value_at_assessment = 95
    db_session.commit()

    client.post(
        f"/api/assessment-facilities/{assessment_facility_id}/run-comparison",
        headers={"Authorization": f"Bearer {manager_token}"},
    )
    return round_id, assessment_facility_id


def test_manager_can_generate_facility_report(
    client, db_session, manager_token, assessor_token, active_facility, active_indicator, seeded_assessor
) -> None:
    _, assessment_facility_id = _prepare_compared_assessment(
        client, db_session, manager_token, assessor_token, active_facility, active_indicator, seeded_assessor
    )
    response = client.post(
        "/api/reports/generate",
        headers={"Authorization": f"Bearer {manager_token}"},
        json={
            "assessment_facility_id": assessment_facility_id,
            "report_type": "FACILITY_DQA_REPORT",
            "include_comments": False,
        },
    )
    assert response.status_code == 200
    assert response.json()["report_type"] == "FACILITY_DQA_REPORT"


def test_manager_can_generate_consolidated_report(
    client, db_session, manager_token, assessor_token, active_facility, active_indicator, seeded_assessor
) -> None:
    round_id, _ = _prepare_compared_assessment(
        client, db_session, manager_token, assessor_token, active_facility, active_indicator, seeded_assessor
    )
    response = client.post(
        "/api/reports/generate",
        headers={"Authorization": f"Bearer {manager_token}"},
        json={
            "assessment_round_id": round_id,
            "report_type": "CONSOLIDATED_UCMB_DQA_REPORT",
            "include_comments": False,
        },
    )
    assert response.status_code == 200
    assert response.json()["report_type"] == "CONSOLIDATED_UCMB_DQA_REPORT"


def test_report_generation_uses_template_fallback_without_ai_api_key(
    client, db_session, manager_token, assessor_token, active_facility, active_indicator, seeded_assessor
) -> None:
    _, assessment_facility_id = _prepare_compared_assessment(
        client, db_session, manager_token, assessor_token, active_facility, active_indicator, seeded_assessor
    )
    response = client.post(
        "/api/reports/generate",
        headers={"Authorization": f"Bearer {manager_token}"},
        json={"assessment_facility_id": assessment_facility_id, "report_type": "FACILITY_DQA_REPORT", "include_comments": False},
    )
    assert response.status_code == 200
    body = response.json()
    assert "Main Findings" in body["generated_content"]
    assert body["structured_input_json"]["source_document_assessment_status"] == "NOT_ASSESSED"
    assert body["structured_input_json"]["finding_blocks"]["findings"][0]["evidence_required_for_closure"]


def test_ai_payload_excludes_comments_by_default(
    client, db_session, manager_token, assessor_token, active_facility, active_indicator, seeded_assessor
) -> None:
    _, assessment_facility_id = _prepare_compared_assessment(
        client, db_session, manager_token, assessor_token, active_facility, active_indicator, seeded_assessor
    )
    response = client.post(
        "/api/reports/generate",
        headers={"Authorization": f"Bearer {manager_token}"},
        json={"assessment_facility_id": assessment_facility_id, "report_type": "FACILITY_DQA_REPORT", "include_comments": False},
    )
    assert response.status_code == 200
    row = response.json()["structured_input_json"]["comparison_rows"][0]
    assert "assessor_comment" not in row


def test_include_comments_true_includes_comments_only_when_requested(
    client, db_session, manager_token, assessor_token, active_facility, active_indicator, seeded_assessor
) -> None:
    _, assessment_facility_id = _prepare_compared_assessment(
        client, db_session, manager_token, assessor_token, active_facility, active_indicator, seeded_assessor
    )
    response = client.post(
        "/api/reports/generate",
        headers={"Authorization": f"Bearer {manager_token}"},
        json={"assessment_facility_id": assessment_facility_id, "report_type": "FACILITY_DQA_REPORT", "include_comments": True},
    )
    assert response.status_code == 200
    row = response.json()["structured_input_json"]["comparison_rows"][0]
    assert row["assessor_comment"] == "Assessor note should be optional in reports."


def test_report_saved_as_generated_not_approved(
    client, db_session, manager_token, assessor_token, active_facility, active_indicator, seeded_assessor
) -> None:
    _, assessment_facility_id = _prepare_compared_assessment(
        client, db_session, manager_token, assessor_token, active_facility, active_indicator, seeded_assessor
    )
    response = client.post(
        "/api/reports/generate",
        headers={"Authorization": f"Bearer {manager_token}"},
        json={"assessment_facility_id": assessment_facility_id, "report_type": "FACILITY_DQA_REPORT", "include_comments": False},
    )
    assert response.status_code == 200
    assert response.json()["status"] == "GENERATED"


def test_manager_can_edit_review_and_approve_report(
    client, db_session, manager_token, assessor_token, active_facility, active_indicator, seeded_assessor
) -> None:
    _, assessment_facility_id = _prepare_compared_assessment(
        client, db_session, manager_token, assessor_token, active_facility, active_indicator, seeded_assessor
    )
    generate_response = client.post(
        "/api/reports/generate",
        headers={"Authorization": f"Bearer {manager_token}"},
        json={"assessment_facility_id": assessment_facility_id, "report_type": "FACILITY_DQA_REPORT", "include_comments": False},
    )
    report_id = generate_response.json()["id"]

    edit_response = client.put(
        f"/api/reports/{report_id}",
        headers={"Authorization": f"Bearer {manager_token}"},
        json={"edited_content": "Edited report body."},
    )
    assert edit_response.status_code == 200
    assert edit_response.json()["edited_content"] == "Edited report body."

    review_response = client.post(f"/api/reports/{report_id}/review", headers={"Authorization": f"Bearer {manager_token}"})
    assert review_response.status_code == 200
    assert review_response.json()["status"] == "REVIEWED"

    approve_response = client.post(f"/api/reports/{report_id}/approve", headers={"Authorization": f"Bearer {manager_token}"})
    assert approve_response.status_code == 200
    assert approve_response.json()["status"] == "APPROVED"


def test_viewer_and_assessor_cannot_generate_official_reports(
    client,
    db_session,
    viewer_token,
    assessor_token,
    manager_token,
    active_facility,
    active_indicator,
    seeded_assessor,
) -> None:
    _, assessment_facility_id = _prepare_compared_assessment(
        client, db_session, manager_token, assessor_token, active_facility, active_indicator, seeded_assessor
    )
    payload = {"assessment_facility_id": assessment_facility_id, "report_type": "FACILITY_DQA_REPORT", "include_comments": False}
    viewer_response = client.post("/api/reports/generate", headers={"Authorization": f"Bearer {viewer_token}"}, json=payload)
    assessor_response = client.post("/api/reports/generate", headers={"Authorization": f"Bearer {assessor_token}"}, json=payload)
    assert viewer_response.status_code == 403
    assert assessor_response.status_code == 403


def test_docx_and_xlsx_export_return_file_responses_and_logs_are_saved(
    client, db_session, manager_token, assessor_token, active_facility, active_indicator, seeded_assessor
) -> None:
    _, assessment_facility_id = _prepare_compared_assessment(
        client, db_session, manager_token, assessor_token, active_facility, active_indicator, seeded_assessor
    )
    generate_response = client.post(
        "/api/reports/generate",
        headers={"Authorization": f"Bearer {manager_token}"},
        json={"assessment_facility_id": assessment_facility_id, "report_type": "FACILITY_DQA_REPORT", "include_comments": False},
    )
    report_id = generate_response.json()["id"]
    client.post(f"/api/reports/{report_id}/approve", headers={"Authorization": f"Bearer {manager_token}"})

    docx_response = client.get(f"/api/reports/{report_id}/export/docx", headers={"Authorization": f"Bearer {manager_token}"})
    xlsx_response = client.get(f"/api/reports/{report_id}/export/xlsx", headers={"Authorization": f"Bearer {manager_token}"})

    assert docx_response.status_code == 200
    assert "attachment" in docx_response.headers["content-disposition"]
    assert xlsx_response.status_code == 200
    assert db_session.query(ExportLog).filter_by(report_id=UUID(report_id)).count() >= 2


def test_generated_report_can_be_exported_to_docx_without_approval(
    client, db_session, manager_token, assessor_token, active_facility, active_indicator, seeded_assessor
) -> None:
    _, assessment_facility_id = _prepare_compared_assessment(
        client, db_session, manager_token, assessor_token, active_facility, active_indicator, seeded_assessor
    )
    generate_response = client.post(
        "/api/reports/generate",
        headers={"Authorization": f"Bearer {manager_token}"},
        json={"assessment_facility_id": assessment_facility_id, "report_type": "FACILITY_DQA_REPORT", "include_comments": False},
    )
    report_id = generate_response.json()["id"]

    docx_response = client.get(f"/api/reports/{report_id}/export/docx", headers={"Authorization": f"Bearer {manager_token}"})

    assert docx_response.status_code == 200
    assert docx_response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    )
    assert docx_response.content.startswith(b"PK")


def test_pdf_export_handles_dependency_availability_gracefully(
    client, db_session, manager_token, assessor_token, active_facility, active_indicator, seeded_assessor
) -> None:
    _, assessment_facility_id = _prepare_compared_assessment(
        client, db_session, manager_token, assessor_token, active_facility, active_indicator, seeded_assessor
    )
    generate_response = client.post(
        "/api/reports/generate",
        headers={"Authorization": f"Bearer {manager_token}"},
        json={"assessment_facility_id": assessment_facility_id, "report_type": "FACILITY_DQA_REPORT", "include_comments": False},
    )
    report_id = generate_response.json()["id"]
    client.post(f"/api/reports/{report_id}/approve", headers={"Authorization": f"Bearer {manager_token}"})
    response = client.get(f"/api/reports/{report_id}/export/pdf", headers={"Authorization": f"Bearer {manager_token}"})
    assert response.status_code in {200, 503}
    if response.status_code == 503:
        assert "PDF export dependency" in response.json()["detail"]


def test_export_requires_authentication(
    client, db_session, manager_token, assessor_token, active_facility, active_indicator, seeded_assessor
) -> None:
    _, assessment_facility_id = _prepare_compared_assessment(
        client, db_session, manager_token, assessor_token, active_facility, active_indicator, seeded_assessor
    )
    generate_response = client.post(
        "/api/reports/generate",
        headers={"Authorization": f"Bearer {manager_token}"},
        json={"assessment_facility_id": assessment_facility_id, "report_type": "FACILITY_DQA_REPORT", "include_comments": False},
    )
    report_id = generate_response.json()["id"]
    client.post(f"/api/reports/{report_id}/approve", headers={"Authorization": f"Bearer {manager_token}"})
    response = client.get(f"/api/reports/{report_id}/export/docx")
    assert response.status_code == 401


def test_report_access_respects_permissions(
    client,
    db_session,
    manager_token,
    assessor_token,
    assessor_two_token,
    active_facility,
    active_indicator,
    seeded_assessor,
) -> None:
    _, assessment_facility_id = _prepare_compared_assessment(
        client, db_session, manager_token, assessor_token, active_facility, active_indicator, seeded_assessor
    )
    generate_response = client.post(
        "/api/reports/generate",
        headers={"Authorization": f"Bearer {manager_token}"},
        json={"assessment_facility_id": assessment_facility_id, "report_type": "FACILITY_DQA_REPORT", "include_comments": False},
    )
    report_id = generate_response.json()["id"]
    allowed = client.get(f"/api/reports/{report_id}", headers={"Authorization": f"Bearer {assessor_token}"})
    denied = client.get(f"/api/reports/{report_id}", headers={"Authorization": f"Bearer {assessor_two_token}"})
    assert allowed.status_code == 200
    assert denied.status_code == 403


def test_ai_generation_log_is_saved(
    client, db_session, manager_token, assessor_token, active_facility, active_indicator, seeded_assessor
) -> None:
    _, assessment_facility_id = _prepare_compared_assessment(
        client, db_session, manager_token, assessor_token, active_facility, active_indicator, seeded_assessor
    )
    response = client.post(
        "/api/reports/generate",
        headers={"Authorization": f"Bearer {manager_token}"},
        json={"assessment_facility_id": assessment_facility_id, "report_type": "FACILITY_DQA_REPORT", "include_comments": False},
    )
    assert response.status_code == 200
    report_id = UUID(response.json()["id"])
    assert db_session.query(AiGenerationLog).filter_by(report_id=report_id).count() == 1


def test_docx_finding_block_export_has_no_narrative_report_section(db_session, seeded_manager) -> None:
    structured = {
        "coverage": {"facilities_assessed": 1, "total_facilities_selected": 1, "administrative_areas_covered": ["Area A"]},
        "dqa_score": {"score_percent": 60, "exact_count": 0, "minor_count": 0, "moderate_count": 0, "major_count": 1, "critical_count": 1, "missing_count": 0},
        "source_document_assessment_status": "NOT_ASSESSED",
        "dhis2_sync_summary": {"response_classification": {"NO_DATA": 1, "TRUE_ZERO": 0, "VALUE_RETURNED": 0, "SYNC_ERROR": 0, "NOT_APPLICABLE": 0, "UNKNOWN": 0}},
        "critical_chase_list": [
            {
                "facility": "Facility A",
                "administrative_area": "Area A",
                "indicator": "Newborn deaths",
                "hmis_code": "105-MA13",
                "register_value": 1,
                "hmis105_value": 1,
                "dhis2_value": None,
                "gap": 0,
                "pattern": "DHIS2 no data",
                "owner_role": "Facility In-charge",
                "proposed_target_date": "Within 30 days of report approval",
                "evidence_required_for_closure": "MPDSR cross-check note.",
            }
        ],
        "finding_blocks": {
            "executive_snapshot": {"headline": "Snapshot", "primary_finding": "Primary.", "management_implication": "Implication.", "urgent_actions": []},
            "scope_and_method": {"scope_summary": "Scope.", "method_summary": "Method.", "denominator_note": "Denominator.", "severity_note": "Severity."},
            "critical_chase_list_intro": "Chase critical rows.",
            "findings": [
                {
                    "finding_number": 1,
                    "finding_title": "DHIS2 no-data responses require separate investigation",
                    "finding_category": "DHIS2 synchronization",
                    "evidence": "One no-data row.",
                    "interpretation": "No-data is not zero.",
                    "affected_facilities": ["Facility A"],
                    "affected_indicators": ["Newborn deaths"],
                    "affected_administrative_areas": ["Area A"],
                    "risk_level": "Critical",
                    "implication": "Requires reconciliation.",
                    "required_action": "Verify DHIS2.",
                    "owner_role": "MoH DHIS2 Team",
                    "proposed_timeline": "Within 30 days",
                    "evidence_required_for_closure": "DHIS2 screenshot.",
                }
            ],
            "dhis2_no_data_review": {"summary": "No-data review.", "interpretation": "Verify.", "required_platform_fix": "Show status."},
            "source_document_review": {"summary": "Source document quality was not fully measured in this round.", "interpretation": "Not assessed.", "next_round_requirement": "Add checklist."},
            "facility_performance_summary": {"summary": "Facility summary."},
            "indicator_performance_summary": {"summary": "Indicator summary."},
            "corrective_action_plan": {"summary": "Action summary.", "actions": []},
            "limitations": ["Not assessed"],
            "next_round_improvements": ["Add checklist"],
            "conclusion": "Conclusion.",
        },
    }
    report = Report(
        report_type=ReportType.CONSOLIDATED_UCMB_DQA_REPORT,
        title="Finding Block Export Test",
        status=ReportStatus.GENERATED,
        generated_content="Finding block content",
        final_content="Finding block content",
        structured_input_json=structured,
        prompt_version="v4-finding-blocks-blended-report",
        generated_by_user_id=seeded_manager.id,
    )
    db_session.add(report)
    db_session.commit()
    db_session.refresh(report)

    _, content, _ = export_report_docx(db_session, report, seeded_manager)
    document = Document(BytesIO(content))
    text = "\n".join(paragraph.text for paragraph in document.paragraphs)
    assert "Narrative Report" not in text
    assert "Executive Snapshot" in text
    assert "Critical Chase List" in text
    assert "Administrative area" in text


def test_xlsx_export_includes_analytical_workbook_sheets(db_session, seeded_manager) -> None:
    structured = {
        "finding_blocks": {"corrective_action_plan": {"actions": []}},
        "dqa_score": {"score_percent": 80, "critical_count": 0},
        "summary": {},
        "dhis2_sync_summary": {"response_classification": {"NO_DATA": 0}},
        "source_document_assessment_status": "NOT_ASSESSED",
        "comparison_rows": [],
        "critical_chase_list": [],
    }
    report = Report(
        report_type=ReportType.EXECUTIVE_SUMMARY,
        title="Analytical Workbook Test",
        status=ReportStatus.APPROVED,
        generated_content="Content",
        final_content="Content",
        structured_input_json=structured,
        prompt_version="v4-finding-blocks-blended-report",
        generated_by_user_id=seeded_manager.id,
    )
    db_session.add(report)
    db_session.commit()
    db_session.refresh(report)

    _, content, _ = export_report_xlsx(db_session, report, seeded_manager)
    workbook = load_workbook(BytesIO(content))
    assert {
        "Executive Dashboard",
        "Facility Ranking",
        "Indicator Ranking",
        "Critical Chase List",
        "Corrective Action Tracker",
        "DHIS2 Sync Audit",
        "Source Document Checklist",
        "Submitted Data",
        "Field Comments",
        "Data Dictionary",
    }.issubset(set(workbook.sheetnames))
